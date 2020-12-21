import csv
import pandas as pd
from openpyxl import load_workbook

wb = load_workbook(filename='c:/SampleData/pandatest.xlsx')

print( wb.sheetnames)
sheet_ranges =  wb['Sheet1']

print(sheet_ranges['A1'].value)
for sheet in wb:
    print(f'----printing first row of sheet {sheet.title}all sheets ----')
    for row in sheet.iter_rows(min_row=1,max_row=2):
        for cell in row:
            print(cell.value)

